#!/usr/bin/env python3
"""
KLG Pre-Bill Audit — deterministic detection engine.

Reads a Clio time-entry export (CSV) for a billing period and flags
entries that draw fee cuts: long vague entries, block billing,
duplicate/near-duplicate descriptions, intra-firm conferencing,
clerical work billed at professional rates, vague conferences/emails,
and several hygiene checks. Produces an .xlsx audit workbook.

Design principle: the script favors RECALL over precision. It
over-flags on purpose. Claude's judgment layer culls false positives
and writes the suggested supplemental descriptions.

Usage:
    pip install openpyxl
    python prebill_audit.py INPUT.csv -o OUTPUT.xlsx --period "May 2026"

CSV column mapping is configured in COLUMN_ALIASES below. Clio exports
vary; add aliases as needed rather than renaming columns by hand.
"""

import argparse
import csv
import re
import sys
from datetime import datetime
from difflib import SequenceMatcher

# ----------------------------------------------------------------------
# CONFIG — tune these to firm policy.
# ----------------------------------------------------------------------

CONFIG = {
    "long_entry_hours": 2.0,
    "outlier_hours": 6.0,
    "thin_description_words": 8,
    "duplicate_similarity": 0.85,
    "duplicate_window_days": 31,
    "billing_increment": 0.1,
    "round_number_share_flag": 0.40,
}

COLUMN_ALIASES = {
    "date":        ["date", "activity date", "service date", "entry date"],
    "timekeeper":  ["user", "timekeeper", "attorney", "name", "person", "billed by"],
    "matter":      ["matter", "matter description", "matter name", "client/matter"],
    "hours":       ["quantity", "hours", "duration", "time", "billed hours"],
    "rate":        ["rate", "billed rate", "hourly rate"],
    "amount":      ["amount", "total", "billed amount", "value"],
    "description": ["description", "narrative", "notes", "activity description"],
}

CLERICAL_TERMS = [
    "e-file", "efile", "e file", "file with", "filing", "uploaded", "upload",
    "download", "bookmark", "format", "formatting", "reformat", "paginate",
    "bates", "scan", "scanning", "photocopy", "copy", "print", "printing",
    "calendar", "calendaring", "diary", "tickle", "organize", "organizing",
    "compile binder", "assemble binder", "tab ", "index the", "save to",
    "save file", "convert to pdf", "create pdf", "load to", "data entry",
    "update the file", "set up file", "open file", "close file",
    "memorandum of costs", "memo of costs", "proof of service", "serve ",
    "schedule ", "reschedule", "confirm receipt", "transmit", "send copy",
]

GENERIC_OPENERS = [
    "research", "review", "prepare", "work on", "working on", "attention to",
    "analysis", "analyze", "draft", "revise", "address", "handle", "deal with",
    "continue", "finish", "complete", "various", "miscellaneous", "general",
]

CONFER_TERMS = [
    "conference", "confer", "meeting", "meet with", "call with", "telephone",
    "phone call", "tc with", "tc ", "discuss", "strategy session",
]
COMM_TERMS = ["email", "e-mail", "correspond", "letter to", "message to", "text to"]

ACTION_VERBS = [
    "research", "review", "draft", "revise", "prepare", "analyze", "edit",
    "call", "email", "confer", "meet", "attend", "argue", "file", "cite-check",
    "outline", "investigate", "negotiate", "respond", "summarize", "organize",
]


# ----------------------------------------------------------------------
# Loading / normalization
# ----------------------------------------------------------------------

def resolve_columns(headers):
    lower = {h.lower().strip(): h for h in headers}
    resolved = {}
    for field, aliases in COLUMN_ALIASES.items():
        for a in aliases:
            if a in lower:
                resolved[field] = lower[a]
                break
    return resolved


def parse_float(v):
    if v is None:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    if s in ("", "-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(v):
    if not v:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%d-%b-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def normalize_desc(text):
    t = (text or "").lower()
    t = re.sub(r"\d{1,2}[/-]\d{1,2}([/-]\d{2,4})?", " ", t)
    t = re.sub(r"[^a-z\s]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def load_entries(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        cols = resolve_columns(headers)
        missing = [k for k in ("date", "timekeeper", "hours", "description") if k not in cols]
        if missing:
            sys.exit(
                f"ERROR: could not map required column(s): {missing}\n"
                f"Found headers: {headers}\n"
                f"Add the right header strings to COLUMN_ALIASES and rerun."
            )
        entries = []
        for i, row in enumerate(reader):
            desc = (row.get(cols["description"], "") or "").strip()
            hours = parse_float(row.get(cols["hours"]))
            rate = parse_float(row.get(cols.get("rate", ""))) if "rate" in cols else None
            amount = parse_float(row.get(cols.get("amount", ""))) if "amount" in cols else None
            if amount is None and hours is not None and rate is not None:
                amount = round(hours * rate, 2)
            entries.append({
                "row": i + 2,
                "date": parse_date(row.get(cols["date"])),
                "timekeeper": (row.get(cols["timekeeper"], "") or "").strip(),
                "matter": (row.get(cols.get("matter", ""), "") or "").strip() if "matter" in cols else "",
                "hours": hours,
                "rate": rate,
                "amount": amount,
                "description": desc,
                "norm": normalize_desc(desc),
                "flags": [],
            })
        return entries


# ----------------------------------------------------------------------
# Detection rules
# ----------------------------------------------------------------------

def add(e, category, severity, note):
    e["flags"].append((category, severity, note))


def contains_any(text, terms):
    return [t for t in terms if t in text]


def _term_regex(term):
    t = term.strip().lower()
    if " " in t or "-" in t:
        return re.compile(r"\b" + re.escape(t), re.I)
    return re.compile(r"\b" + re.escape(t) + r"\w*", re.I)


_CLERICAL_RES = [_term_regex(t) for t in CLERICAL_TERMS]


def rule_clerical(e):
    raw = (e["description"] or "").lower()
    raw_hits = [t for t, rx in zip(CLERICAL_TERMS, _CLERICAL_RES) if rx.search(raw)]
    if raw_hits:
        add(e, "Clerical / non-billable", "High",
            "Reads as clerical/secretarial work. Not separately compensable at "
            "a professional rate. Write off, reassign to overhead, or confirm it "
            "is genuinely substantive.")


def rule_long_vague(e):
    if e["hours"] is None:
        return
    if e["hours"] >= CONFIG["long_entry_hours"]:
        words = len((e["description"] or "").split())
        first = e["norm"].split()[:1]
        generic = bool(first) and first[0] in [normalize_desc(g).split()[0] for g in GENERIC_OPENERS if g]
        if words < CONFIG["thin_description_words"] or (generic and words < 14):
            add(e, "Long entry, thin description", "High",
                f"{e['hours']}h on a {words}-word description. A reviewing court "
                "cannot test reasonableness of a long block this vague. Itemize "
                "the discrete tasks and what each produced.")


def rule_block_billing(e):
    raw = (e["description"] or "").lower()
    distinct_verbs = {v for v in ACTION_VERBS if re.search(rf"\b{re.escape(v)}\b", raw)}
    semicolons = raw.count(";")
    if len(distinct_verbs) >= 2 or semicolons >= 1:
        sev = "High" if (e["hours"] or 0) >= CONFIG["long_entry_hours"] else "Medium"
        add(e, "Block billing", sev,
            "Multiple discrete tasks lumped in one entry"
            + (f" ({', '.join(sorted(distinct_verbs))})" if distinct_verbs else "")
            + ". Split into separate entries with task-level time.")


def rule_vague_comm(e):
    raw = (e["description"] or "").lower()
    is_conf = bool(contains_any(raw, CONFER_TERMS))
    is_comm = bool(contains_any(raw, COMM_TERMS))
    if is_conf or is_comm:
        has_subject = bool(re.search(r"\b(re|regarding|about|concerning)\b", raw))
        if not has_subject and len((e["description"] or "").split()) < 8:
            add(e, "Conference/email without subject", "Medium",
                "Communication entry with no stated subject matter. Add the topic "
                "('re ...') so the entry shows what was actually discussed.")


def rule_outlier(e):
    if e["hours"] is not None and e["hours"] >= CONFIG["outlier_hours"]:
        add(e, "Large single block — billing judgment", "Medium",
            f"{e['hours']}h in one entry. Confirm this reflects genuine continuous "
            "work and exercise billing judgment before it goes out.")


def rule_increment(e):
    if e["hours"] is None:
        return
    inc = CONFIG["billing_increment"]
    rem = round((e["hours"] / inc) % 1, 6)
    if rem not in (0.0,) and abs(rem - 1.0) > 1e-6:
        add(e, "Off-increment entry", "Low",
            f"{e['hours']}h is not a clean multiple of {inc}h.")


def rule_hygiene(e):
    if not e["description"]:
        add(e, "Missing description", "High", "Empty narrative — cannot be billed.")
    if e["hours"] is None or e["hours"] <= 0:
        add(e, "Bad duration", "High", "Zero, missing, or negative hours.")
    if e["date"] is None:
        add(e, "Unparseable date", "Low", "Date did not parse — check the export.")


def rule_cross_matter(e, _):
    raw = e["description"] or ""
    case_nos = re.findall(r"\b[A-Z]\d{6}\b", raw)
    if case_nos and e["matter"]:
        if not any(cn in e["matter"] for cn in case_nos):
            add(e, "Possible cross-matter reference", "Medium",
                f"Narrative cites case number(s) {', '.join(case_nos)} not in this "
                "matter's name. Confirm the time was booked to the right matter.")


def detect_duplicates(entries):
    by_tk = {}
    for e in entries:
        if e["norm"]:
            by_tk.setdefault(e["timekeeper"], []).append(e)
    for tk, group in by_tk.items():
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                a, b = group[i], group[j]
                if a["date"] and b["date"]:
                    if abs((a["date"] - b["date"]).days) > CONFIG["duplicate_window_days"]:
                        continue
                ratio = SequenceMatcher(None, a["norm"], b["norm"]).ratio()
                if ratio >= CONFIG["duplicate_similarity"]:
                    add(a, "Duplicate / near-duplicate", "High",
                        f"Near-identical to entry on row {b['row']} "
                        f"({b['date']}, {b['hours']}h) — similarity {ratio:.2f}. "
                        "Confirm these are distinct tasks, not the same work "
                        "described twice.")


def detect_intrafirm_conferences(entries):
    conf = [e for e in entries if contains_any((e["description"] or "").lower(), CONFER_TERMS)]
    by_date = {}
    for e in conf:
        if e["date"]:
            by_date.setdefault(e["date"], []).append(e)
    for date, group in by_date.items():
        tks = {e["timekeeper"] for e in group}
        if len(tks) >= 2:
            for i in range(len(group)):
                for j in range(i + 1, len(group)):
                    a, b = group[i], group[j]
                    if a["timekeeper"] == b["timekeeper"]:
                        continue
                    overlap = SequenceMatcher(None, a["norm"], b["norm"]).ratio()
                    if overlap >= 0.45:
                        total = sum(x["hours"] or 0 for x in (a, b))
                        add(a, "Intra-firm conference", "Medium",
                            f"{a['timekeeper']} and {b['timekeeper']} both billed a "
                            f"conference on {date} (combined {total:.1f}h). Confirm "
                            "multiple billers were necessary.")


def firm_level_notes(entries):
    notes = []
    timed = [e for e in entries if e["hours"] is not None]
    if timed:
        round_ct = sum(1 for e in timed if abs((e["hours"] * 2) - round(e["hours"] * 2)) < 1e-6)
        share = round_ct / len(timed)
        if share > CONFIG["round_number_share_flag"]:
            notes.append(
                f"{share:.0%} of entries land on whole/half-hour values. A bill that "
                "reads as estimated rather than contemporaneous invites a haircut. "
                "Confirm timers are running in real time."
            )
    return notes


def audit(entries):
    for e in entries:
        rule_hygiene(e)
        rule_clerical(e)
        rule_long_vague(e)
        rule_block_billing(e)
        rule_vague_comm(e)
        rule_outlier(e)
        rule_increment(e)
        rule_cross_matter(e, None)
    detect_duplicates(entries)
    detect_intrafirm_conferences(entries)
    return entries


# ----------------------------------------------------------------------
# Output
# ----------------------------------------------------------------------

SEV_ORDER = {"High": 0, "Medium": 1, "Low": 2}


def write_xlsx(entries, firm_notes, out_path, period=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True)
    sev_fill = {
        "High": PatternFill("solid", fgColor="F8CBAD"),
        "Medium": PatternFill("solid", fgColor="FFE699"),
        "Low": PatternFill("solid", fgColor="D9E1F2"),
    }
    wrap = Alignment(wrap_text=True, vertical="top")

    flagged = [e for e in entries if e["flags"]]

    def entry_severity(e):
        return min((SEV_ORDER[f[1]] for f in e["flags"]), default=3)

    flagged.sort(key=lambda e: (entry_severity(e), -(e["hours"] or 0)))

    ws = wb.active
    ws.title = "Flagged Entries"
    headers = ["Severity", "Date", "Timekeeper", "Matter", "Hours", "Rate",
               "Amount", "Original description", "Issues flagged",
               "Suggested fix / supplemented language", "Resolution"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).fill = hdr_fill
        ws.cell(1, c).font = hdr_font
    for e in flagged:
        top = min(e["flags"], key=lambda f: SEV_ORDER[f[1]])[1]
        notes = "\n".join(f"• [{f[1]}] {f[0]}: {f[2]}" for f in
                          sorted(e["flags"], key=lambda f: SEV_ORDER[f[1]]))
        ws.append([
            top, str(e["date"] or ""), e["timekeeper"], e["matter"],
            e["hours"], e["rate"], e["amount"], e["description"], notes, "", "",
        ])
        r = ws.max_row
        ws.cell(r, 1).fill = sev_fill.get(top, sev_fill["Low"])
        for col in (8, 9, 10):
            ws.cell(r, col).alignment = wrap
    widths = [10, 12, 14, 22, 7, 7, 10, 50, 50, 40, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    s = wb.create_sheet("Summary", 0)
    s.append(["KLG Pre-Bill Audit — Summary"])
    s["A1"].font = Font(bold=True, size=14)
    if period:
        s.append([f"Billing period: {period}"])
    s.append([f"Total entries: {len(entries)}"])
    s.append([f"Flagged entries: {len(flagged)}"])
    flagged_hours = sum(e["hours"] or 0 for e in flagged)
    flagged_amt = sum(e["amount"] or 0 for e in flagged)
    s.append([f"Flagged hours: {flagged_hours:.1f}"])
    s.append([f"Flagged amount: ${flagged_amt:,.2f}"])
    s.append([])

    cat_count, cat_hours = {}, {}
    for e in flagged:
        for cat in {f[0] for f in e["flags"]}:
            cat_count[cat] = cat_count.get(cat, 0) + 1
            cat_hours[cat] = cat_hours.get(cat, 0) + (e["hours"] or 0)
    s.append(["By issue category", "Entries", "Hours"])
    hr = s.max_row
    for c in range(1, 4):
        s.cell(hr, c).fill = hdr_fill
        s.cell(hr, c).font = hdr_font
    for cat in sorted(cat_count, key=lambda k: -cat_count[k]):
        s.append([cat, cat_count[cat], round(cat_hours[cat], 1)])
    s.append([])

    tk_count, tk_hours = {}, {}
    for e in flagged:
        tk_count[e["timekeeper"]] = tk_count.get(e["timekeeper"], 0) + 1
        tk_hours[e["timekeeper"]] = tk_hours.get(e["timekeeper"], 0) + (e["hours"] or 0)
    s.append(["By timekeeper", "Flagged entries", "Flagged hours"])
    hr = s.max_row
    for c in range(1, 4):
        s.cell(hr, c).fill = hdr_fill
        s.cell(hr, c).font = hdr_font
    for tk in sorted(tk_count, key=lambda k: -tk_hours[k]):
        s.append([tk or "(blank)", tk_count[tk], round(tk_hours[tk], 1)])
    s.append([])

    if firm_notes:
        s.append(["Firm-level notes"])
        s.cell(s.max_row, 1).font = Font(bold=True)
        for n in firm_notes:
            s.append([n])
            s.cell(s.max_row, 1).alignment = wrap
    for col, w in (("A", 40), ("B", 16), ("C", 16)):
        s.column_dimensions[col].width = w

    wb.save(out_path)
    return len(flagged), flagged_hours, flagged_amt


def main():
    ap = argparse.ArgumentParser(description="KLG pre-bill audit")
    ap.add_argument("input", help="Clio time-entry CSV export")
    ap.add_argument("-o", "--output", default="prebill_audit.xlsx")
    ap.add_argument("--period", default=None, help="Label, e.g. 'May 2026'")
    args = ap.parse_args()

    entries = load_entries(args.input)
    audit(entries)
    firm_notes = firm_level_notes(entries)
    n, hrs, amt = write_xlsx(entries, firm_notes, args.output, args.period)
    print(f"Audited {len(entries)} entries — {n} flagged "
          f"({hrs:.1f} h, ${amt:,.2f}). Workbook: {args.output}")
    if firm_notes:
        print("Firm-level notes:")
        for note in firm_notes:
            print("  - " + note)


if __name__ == "__main__":
    main()
